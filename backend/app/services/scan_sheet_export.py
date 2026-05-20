"""Excel export for scan-sheet downloads.

Loads the committed TEMPLATE.xlsx (preserving header layout, navy/blue
column header strip, conditional-formatting dup detection on Serial and
IMEI columns) and writes the receipt's header + scan rows into the
existing cells.

Two flavors:
  build_single_container_workbook(detail) → BytesIO  (one container)
  build_bulk_workbook(details)            → BytesIO  (one sheet per container)
"""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.scan_sheet import AuditSheetDetail, ScanRow, ScanSheetHeader

TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "templates" / "scan_sheet_template.xlsx"
)

# Excel sheet-name limit + forbidden characters
_FORBIDDEN_SHEET_CHARS = set(r"[]:*?/\\")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Sanitize a container_no into a valid Excel sheet name, deduping if
    the same number appears twice (shouldn't happen but cheap to guard)."""
    clean = "".join("_" if ch in _FORBIDDEN_SHEET_CHARS else ch for ch in name)
    clean = clean[:31] or "Sheet"
    candidate = clean
    i = 2
    while candidate in used:
        suffix = f"_{i}"
        candidate = clean[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


# ─── Cell map (from TEMPLATE.xlsx inspection) ───────────────────────────
#
#   B3        Received Date            (peach input cell)
#   F3:G3     3PL Location             — already pre-filled in template
#   B5        Container Number         (peach input cell)
#   F5        BOL # or Tracking #      (light peach hint cell)
#   B6        Start Timestamp          (value cell to the right of A6 label)
#   D6        Completed Timestamp      (value cell to the right of C6 label)
#   E6        Completion checkbox      (yellow boolean cell)
#   Rows 8+   Scan rows: A=container_no  B=sku  C=qty  D=serial
#             E=imei  F=scanned_by  G=notes
#
# Conditional-formatting dup-detection rules on D and E columns carry
# over automatically — we just have to write values into those columns.


def _fill_header(ws: Worksheet, header: ScanSheetHeader) -> None:
    ws["B3"] = header.received_date
    ws["B3"].number_format = "yyyy-mm-dd"
    ws["B5"] = header.container_no
    ws["F5"] = header.bol_number or ""
    ws["B6"] = header.start_timestamp.replace(tzinfo=None)
    ws["B6"].number_format = "yyyy-mm-dd hh:mm:ss"
    if header.completed_timestamp is not None:
        ws["D6"] = header.completed_timestamp.replace(tzinfo=None)
        ws["D6"].number_format = "yyyy-mm-dd hh:mm:ss"
    else:
        ws["D6"] = None
    ws["E6"] = bool(header.is_completed)


def _fill_rows(ws: Worksheet, container_no: str, rows: list[ScanRow]) -> None:
    """Append scan rows starting at row 8. Row 8 in the template is an
    example placeholder — we clobber it with real data (or blanks)."""
    # First clear out any example values in rows 8..N (the template
    # ships with E.g. SEKU... in row 8).
    start = 8
    # The template's data area extends ~327 rows for column D conditional
    # formatting; we'll only overwrite as many rows as we have data for,
    # leaving the rest as the template's blank styled cells.
    for offset, r in enumerate(rows):
        ridx = start + offset
        ws.cell(row=ridx, column=1, value=container_no)              # A
        ws.cell(row=ridx, column=2, value=r.sku)                     # B
        ws.cell(row=ridx, column=3, value=r.qty)                     # C
        ws.cell(row=ridx, column=4, value=r.serial_number)           # D
        ws.cell(row=ridx, column=5, value=r.imei)                    # E (always None today)
        ws.cell(row=ridx, column=6, value=r.scanned_by)              # F
        ws.cell(row=ridx, column=7, value=r.notes)                   # G

    # If no rows written, clear the example row 8 so the operator's
    # download isn't confused with placeholder text.
    if not rows:
        for col in range(1, 8):
            ws.cell(row=8, column=col, value=None)


# ─── Public API ─────────────────────────────────────────────────────────


def build_single_container_workbook(detail: AuditSheetDetail) -> io.BytesIO:
    """Return an in-memory .xlsx for a single container — TEMPLATE.xlsx
    clone with the header + scan rows filled in."""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    _fill_header(ws, detail.header)
    _fill_rows(ws, detail.header.container_no, detail.rows)
    # Rename the sole sheet to the container number for clarity when the
    # auditor opens the file outside our portal.
    ws.title = _safe_sheet_name(detail.header.container_no, set())
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_bulk_workbook(details: list[AuditSheetDetail]) -> io.BytesIO:
    """Return an in-memory .xlsx with one TEMPLATE-style sheet per container.

    Approach: load the template, fill the first sheet with details[0],
    then for each additional container copy that sheet within the same
    workbook (which preserves all styling + conditional formatting),
    rename it, and fill it with the next container's data.
    """
    if not details:
        raise ValueError("build_bulk_workbook requires at least one detail")

    wb = load_workbook(TEMPLATE_PATH)
    base_ws = wb.active

    used_names: set[str] = set()

    # First container goes into the template's existing sheet.
    first = details[0]
    _fill_header(base_ws, first.header)
    _fill_rows(base_ws, first.header.container_no, first.rows)
    base_ws.title = _safe_sheet_name(first.header.container_no, used_names)

    # Remaining containers: copy_worksheet preserves cell values, styles,
    # merged ranges, column widths, and conditional formatting that were
    # set on the *template* tab. We then OVERWRITE the data cells (B3,
    # B5, F5, B6, D6, E6 + rows 8+) with this container's specifics.
    for d in details[1:]:
        new_ws = wb.copy_worksheet(base_ws)
        new_ws.title = _safe_sheet_name(d.header.container_no, used_names)
        # Clear the previous container's data rows (8..max) so we don't
        # carry forward rows from the first container into this sheet.
        for r in range(8, new_ws.max_row + 1):
            for c in range(1, 8):
                new_ws.cell(row=r, column=c, value=None)
        _fill_header(new_ws, d.header)
        _fill_rows(new_ws, d.header.container_no, d.rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
